"""
One-off generator for the RiftRecall Card Recall Master Sheet -- a
comprehensive, one-row-per-(card, question type) export covering every
currently eligible attribute question across the full card set.

Not meant to be re-run on a schedule; it's a snapshot-generation tool, paired
with scripts/apply_master_sheet.py which reads the sheet back in.
"""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

MODE_LABELS = {
    "energyCost": "Energy Cost",
    "powerCost": "Power Cost",
    "might": "Might",
    "keyword": "Keyword",
    "speed": "Speed",
    "name": "Name",
    "text": "Text",
    "fillBlank": "Fill in the Blanks",
}

cards = {c["id"]: c for c in json.load(open("src/data/cards.json"))}
baseline_rows = json.load(open("/tmp/baseline_rows.json"))
existing_questions = json.load(open("src/data/quizQuestions.json"))
existing_overrides = json.load(open("src/data/quizOverrides.json"))

HEADERS = [
    "Card Code",
    "Card Name \n(ref. only)",
    "Set \n(ref. only)",
    "Type \n(ref. only)",
    "Subtype (ref. only)",
    "Category",
    "Enabled",
    "Current Auto Prompt \n(ref. only)",
    "Current Auto Correct Answer \n(ref. only)",
    "Current Auto Distractors \n(ref. only)",
    "Your Custom Prompt",
    "Your Custom Correct Answer",
    "Your Distractor 1",
    "Your Distractor 2",
    "Your Distractor 3",
    "Your Distractor 4",
    "Your Distractor 5",
    "Needs Fixing",
    "Notes / Context",
]

# 1-indexed column roles, matching Ashwin's standardized sheet formatting:
#  - REFERENCE_COLS: the read-only "(ref. only)" snapshot columns (grey body)
#  - BLUE_HEADER_COLS: columns whose HEADER is blue (#4A86E8) = ref/derived
#    (the Set/Type/Subtype metadata + the three Current Auto snapshot cols)
#  - all other headers are navy (#2B3A55) = key/editable columns
REFERENCE_COLS = {8, 9, 10}
BLUE_HEADER_COLS = {3, 4, 5, 8, 9, 10}
# Distractor slots are columns 13-17 (five, not six). Needs Fixing = col 18,
# Notes = col 19.

rows = []
for r in baseline_rows:
    card = cards[r["id"]]
    mode = r["mode"]
    label = MODE_LABELS[mode]

    correct = r["options"][r["correctIndex"]]
    distractors = [o for i, o in enumerate(r["options"]) if i != r["correctIndex"]]
    distractor_str = " | ".join(distractors)

    # For fill-in-the-blank, the prompt alone ("Fill in the blanks:") doesn't
    # show WHAT is being asked — the blanked sentence is in the caption. Fold
    # it into the reference prompt so the sheet is reviewable at a glance.
    ref_prompt = r["prompt"]
    if mode == "fillBlank" and r.get("caption"):
        ref_prompt = f'{r["prompt"]}  [{r["caption"]}]'

    # Auto-flag fill-in-the-blank questions whose two blanks are different
    # KINDS of value (e.g. a recycle cost + a Might buff) — those read less
    # cleanly than two same-kind numbers ("discard N, draw N") and are worth
    # a manual look. Same-kind and non-fillBlank rows are left unflagged.
    needs_fixing = ""
    notes = ""
    if mode == "fillBlank" and r.get("multiKind"):
        needs_fixing = "TRUE"
        kinds = r.get("kinds") or []
        notes = f"Auto-flagged: two blanks are different kinds ({' + '.join(kinds)}) — review wording."

    override = existing_overrides.get(card["id"], {}).get(mode)
    enabled = "FALSE" if override is False else ("TRUE" if override is True else "")

    custom_variants = existing_questions.get(card["id"], {}).get(mode, [])
    if custom_variants:
        for variant in custom_variants:
            pool = variant.get("distractorPool", [])
            padded = pool + [""] * (5 - len(pool))
            rows.append([
                card["id"], card["name"], card["setId"], card["type"], card.get("subtype") or "",
                label, enabled,
                ref_prompt, correct, distractor_str,
                variant["prompt"], variant["correctAnswer"],
                *padded[:5],
                needs_fixing, notes,
            ])
    else:
        rows.append([
            card["id"], card["name"], card["setId"], card["type"], card.get("subtype") or "",
            label, enabled,
            ref_prompt, correct, distractor_str,
            "", "", "", "", "", "", "",
            needs_fixing, notes,
        ])

# Sort for easy browsing: by Card Code, then Category
CATEGORY_ORDER = [
    "Energy Cost", "Power Cost", "Might", "Keyword", "Speed", "Name", "Text", "Fill in the Blanks",
]
rows.sort(key=lambda row: (row[0], CATEGORY_ORDER.index(row[5])))

print(f"Total rows: {len(rows)}")
custom_count = sum(1 for row in rows if row[10])
print(f"Rows pre-filled with an existing custom variant: {custom_count}")

# ---------------------------------------------------------------- workbook

wb = openpyxl.Workbook()

# Header palette matches Ashwin's standardized sheet:
#   navy  #2B3A55 = key / editable columns
#   blue  #4A86E8 = reference / derived columns (metadata + Current Auto)
NAVY = "FF2B3A55"
BLUE = "FF4A86E8"
header_font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
navy_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
blue_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
ref_fill = PatternFill(start_color="FFEDEDED", end_color="FFEDEDED", fill_type="solid")
ref_font = Font(name="Arial", size=10, italic=True, color="595959")
normal_font = Font(name="Arial", size=10)

# --- Data sheet FIRST (active), named exactly as Ashwin's tab ---
ws = wb.active
ws.title = "RA_Card Questions Control Sheet"

for col_idx, header in enumerate(HEADERS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = blue_fill if col_idx in BLUE_HEADER_COLS else navy_fill
    cell.alignment = Alignment(wrap_text=True, vertical="center")

for row_idx, row_data in enumerate(rows, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        if col_idx in REFERENCE_COLS:
            cell.font = ref_font
            cell.fill = ref_fill
        else:
            cell.font = normal_font

ws.freeze_panes = "A2"
last_col = get_column_letter(len(HEADERS))
ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

# Column widths matched to Ashwin's sheet (A..S).
widths = [
    12.25, 22.75, 11.38, 11.63, 13.25, 12.5, 10.75, 35.0, 27.5, 35.0,
    62.13, 16.25, 13.75, 13.0, 13.0, 13.0, 13.0, 10.5, 26.25,
]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# TRUE/FALSE dropdowns on Enabled (col G) and Needs Fixing (col R = 18).
enabled_dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
ws.add_data_validation(enabled_dv)
enabled_dv.add(f"G2:G{len(rows) + 1}")

needs_fixing_dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
ws.add_data_validation(needs_fixing_dv)
needs_fixing_dv.add(f"R2:R{len(rows) + 1}")

# --- Read Me tab SECOND (reference; safe to ignore or delete on copy) ---
legend = wb.create_sheet("Read Me First")
legend_font = Font(name="Arial", size=11)
legend_bold = Font(name="Arial", size=11, bold=True)
legend_title = Font(name="Arial", size=14, bold=True)

legend_lines = [
    ("RA_Card Questions Control Sheet", legend_title),
    ("", legend_font),
    ("One row = one (card, question type) combination currently eligible in RiftRecall.", legend_font),
    ("", legend_font),
    ("COLUMNS", legend_bold),
    ("Card Code / Card Name / Set / Type / Subtype", legend_font),
    ("  Reference only. Not read on import.", legend_font),
    ("Category", legend_font),
    ("  Energy Cost, Power Cost, Might, Keyword, Speed, Name, Text, or Fill in the Blanks.", legend_font),
    ("Enabled", legend_font),
    ("  Blank = keep the app's normal auto-eligibility. FALSE = turn this question off for", legend_font),
    ("  this card. TRUE = force it on (only works if the card data supports it).", legend_font),
    ("Current Auto Prompt / Correct Answer / Distractors (ref. only, grey)", legend_font),
    ("  A real snapshot of what the app shows now if you leave the Custom columns blank.", legend_font),
    ("  Editing these grey columns does NOTHING on import. For Fill in the Blanks, the", legend_font),
    ("  blanked sentence is shown in [brackets] after the prompt.", legend_font),
    ("Your Custom Prompt / Your Custom Correct Answer / Your Distractor 1-5", legend_font),
    ("  Leave ALL blank to keep live auto-generation. Fill in Prompt + Correct Answer to PIN", legend_font),
    ("  this exact question. Provide 2 distractors for a 3-answer (1x3) question or 3+ for a", legend_font),
    ("  4-answer (2x2) question; the app samples 3 from your pool each time it's shown.", legend_font),
    ("  A distractor equal to the correct answer, or a duplicate, is dropped automatically.", legend_font),
    ("Needs Fixing / Notes / Context", legend_font),
    ("  Your triage columns. Never read on import. Fill-in-the-blank rows whose two blanks are", legend_font),
    ("  different KINDS of value are auto-flagged here for review.", legend_font),
    ("", legend_font),
    ("HOW IMPORT WORKS", legend_bold),
    ("Re-export the RA_Card Questions Control Sheet tab as CSV, then run:", legend_font),
    ("  python3 scripts/apply_master_sheet.py path/to/exported_sheet.csv", legend_font),
    ("This writes src/data/quizOverrides.json (from Enabled) and quizQuestions.json (from any", legend_font),
    ("row with a non-blank Your Custom Prompt + Correct Answer). Blank-custom rows are left to", legend_font),
    ("live auto-generation, so nothing you didn't touch gets frozen.", legend_font),
]
for i, (text, font) in enumerate(legend_lines, start=1):
    cell = legend.cell(row=i, column=1, value=text)
    cell.font = font
legend.column_dimensions["A"].width = 105

wb.save("/mnt/user-data/outputs/RA_Card Questions Control Sheet.xlsx")
print("Wrote RA_Card Questions Control Sheet.xlsx")
