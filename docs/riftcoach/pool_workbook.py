#!/usr/bin/env python3
"""
Pre-Rift pool workbook generator (xlsx).

CANONICAL FILE. Overwrite and re-commit; git history is the version log.

Sheet format Ashwin asked for: NO frozen panes, and the input columns
(In deck? / Copies used / Notes) on the RIGHT of the looked-up data.

BUG FIXED 2026-07-30: the turn-1 counter used ISERROR(SEARCH("Poro", ...)) to
exclude Ol' Poro and silently excluded Punching Poro too, undercounting turn-1
units in both the python summary and the delivered spreadsheet formula. Now an
exact-name set, BARRED_FROM_TURN1. Match card names on exact name or collector
number, never substring.

BUG FIXED 2026-07-30 (repo-wide audit): find() carried the same substring risk
as a general startswith() fallback used to resolve every pool/deck lookup, not
just the turn-1 counter -- same class of bug, just not yet triggered by
today's data. Removed; find() is now exact-name-only.
"""

import csv, re
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------- pool as read off photos 1-7 (photo 8 = the built deck)
POOL = [
    # 0847 Legends / Battlefields / Signature
    ("Akali, Rogue Assassin",1,"1"), ("Mel, Soul's Reflection",1,"1"),
    ("Jayce, Defender of Tomorrow",1,"1"), ("Public Execution",1,"1"),
    ("Risen Altar",1,"1"), ("Mystic Vortex",1,"1"), ("Shadow Temple",1,"1"),
    ("Heisho, Shell of the World",1,"1"),
    # 0846 Calm
    ("Serene Ascetic",1,"2"), ("Pakaa Protector",2,"2"), ("Sandstone Chimera",1,"2"),
    ("Ol' Poro",1,"2"), ("Field Musicians",2,"2"), ("Resonating Strike",1,"2"),
    ("Mournful Witness",1,"2"), ("Akali, Silent",1,"2"), ("Affectionate Poro",1,"2"),
    ("Esteemed Hierophant",1,"2"),
    # 0845 Mind
    ("Questionable Tome",1,"3"), ("Cloud Drake",1,"3"), ("Apprentice Mage",1,"3"),
    ("Clairvoyance",1,"3"), ("Iterative Design",1,"3"), ("Covert Informant",1,"3"),
    ("Decree of Insight",1,"3"), ("Otterpus",1,"3"), ("Shock Blast",1,"3"),
    ("Dredge Up",1,"3"), ("Mesmerize",1,"3"),
    # 0844 Chaos
    ("Decree of Discord",1,"4"), ("Mask Mother",2,"4"), ("Kinkou Lifeblade",1,"4"),
    ("Ravenbloom Prefect",1,"4"), ("Wind and Ghosts",1,"4"), ("Shadow Order Disciple",2,"4"),
    ("Zed, Without a Sound",1,"4"), ("Twilight Step",1,"4"), ("Tornado Warrior",1,"4"),
    ("Gust Monk",1,"4"), ("Shadowblade Lurker",2,"4"), ("Spiderling",1,"4"),
    ("Stargazer",1,"4"), ("Up from the Deep",1,"4"), ("Shadows of the Past",1,"4"),
    # 0843 Order
    ("Escaped Grayback",1,"5"), ("Soulspinner",1,"5"), ("Sacred Protector",2,"5"),
    ("Horns of the Dragon",1,"5"), ("Aurok General",1,"5"), ("Solari Sunhawk",1,"5"),
    ("Disciple of Shen",1,"5"), ("Masa, Crashing Thunder",1,"5"), ("Dragon Form",1,"5"),
    ("Noxian Emissary",1,"5"), ("Fallen Feline",1,"5"), ("Reluctant Leader",1,"5"),
    # 0842 Body
    ("Platewyrm Egg",3,"6"), ("Ambessa, The Wolf",1,"6"), ("Rampage",1,"6"),
    ("Brutal Hunter",1,"6"), ("Guttural Roar",1,"6"), ("Fretful Feline",1,"6"),
    ("Repair Specialist",1,"6"), ("Legion Marauder",1,"6"), ("Noxian Demolitionist",1,"6"),
    # 0841 Fury
    ("Decree of Rage",2,"7"), ("Baccai Reaper",2,"7"), ("Oasis Raider",1,"7"),
    ("Punching Poro",3,"7"), ("Blade Twirler",2,"7"), ("Pendulum Blade",1,"7"),
    ("Perfect Execution",1,"7"), ("Consuming Curse",1,"7"), ("Forsaken Baccai",2,"7"),
    ("Dune Surfer",1,"7"), ("Ruthless Strike",1,"7"), ("Brittle Steel",1,"7"),
    ("Zed, Master of Shadows",1,"7"), ("Baccai Sandspinner",1,"7"), ("Endless Riches",1,"7"),
]

# ---------------- photo 8: the deck as built
DECK = {
    "Akali, Rogue Assassin":1, "Mystic Vortex":1, "Risen Altar":1,
    "Akali, Silent":1, "Kinkou Lifeblade":1, "Esteemed Hierophant":1,
    "Mournful Witness":1, "Ol' Poro":1, "Pakaa Protector":2, "Ruthless Strike":1,
    "Consuming Curse":1, "Blade Twirler":2, "Tornado Warrior":1, "Punching Poro":2,
    "Pendulum Blade":1, "Perfect Execution":1, "Ravenbloom Prefect":1, "Twilight Step":1,
    "Shadow Order Disciple":2, "Wind and Ghosts":1, "Zed, Without a Sound":1,
    "Affectionate Poro":1, "Dune Surfer":1, "Serene Ascetic":1, "Forsaken Baccai":1,
}

inv = [x for x in csv.DictReader(
    open("/mnt/user-data/uploads/Riftbound_stuff_-_Master_Card_Inventory__1_.csv",
         encoding="utf-8-sig")) if x["Set"] == "Vendetta"]
norm = lambda s: re.sub(r"[^a-z0-9 ]", "", s.lower().replace("&", "and").replace("\u2019", "'")).strip()
BY = {norm(r["Card Name"]): r for r in inv}


def find(n):
    k = norm(n)
    return BY.get(k)


V7 = {
 "Fury": (["Punching Poro","Shadow Fiend","Baccai Reaper","Perfect Execution","Ruthless Strike","Oasis Raider","Morgana Vindictive"],
          ["Consuming Curse","Forsaken Baccai","Akali Deadly Weapon","Dune Surfer","Pendulum Blade","Twilight Reveler","Blade Twirler","Zed From the Shadows","Renekton Rage Fueled","Eclipse Dragon"],
          ["Rage Amplifier","Shadow Assassin","Baccai Sandspinner"],
          ["Brittle Steel","Endless Riches","Decree of Rage"]),
 "Calm": (["Twilight Shroud","Mournful Witness","Ol' Poro","Resonating Strike","Serene Ascetic","Akali Silent","Pakaa Protector"],
          ["Hand Hammer","Affectionate Poro","Sanction","Field Musicians","Tomb-Raider Barbara","Esteemed Hierophant","Shen Scourge of Shadows","Astral Heron","Nasus Ascended"],
          ["Frostcoat Mother","Riven Shattered","Sandstone Chimera"],
          ["Steel Paws","Crumbling Sands","Helm of Suppression","Decree of Focus"]),
 "Mind": (["Mesmerize","Apprentice Mage","Covert Informant","Shock Blast","Mel Newly Awakened","Nasus Guardian of Knowledge"],
          ["Otterpus","Dredge Up","Applied Researchers","Grumpy Rockbear","Iterative Design","Cloud Drake","Jayce Brilliant Inventor","Swain Visionary"],
          ["Patched Porobot","Temporal Breach","Questionable Tome","Sky Cruiser","Plaza Guardian"],
          ["Bottled Constellation","Clairvoyance","Decree of Insight"]),
 "Body": (["Guttural Roar","Legion Marauder","Brutal Hunter","Rampage","Ambessa The Wolf","Jayce Hammer in Hand","Onslaught"],
          ["Jagged Cutlass","Baccai Witherclaw","Profiteer","Tools of Empire","Dame the Despoiler","Gangplank Naval"],
          ["Repair Specialist","Hextech Disc","Renekton Brute","Wild Claw","Cataclysmic Duel","Corrupted Dragon"],
          ["Noxian Demolitionist","Platewyrm Egg","Decree of Strength"]),
 "Chaos": (["Gust Monk","Shadow Order Disciple","Kennen Storm of Shuriken","Tornado Warrior","Wind and Ghosts","Tail-Cloaked Matriarch"],
           ["Twilight Step","Mask Mother","Ravenbloom Prefect","Kinkou Lifeblade","Mel Defiant Soul","Zed Without a Sound","Illaoi","Minah Swiftfoot","Ocean Drake"],
           ["Up from the Deep","Shadows of the Past","Shadowblade Lurker","Stargazer","Forgotten Relic","Kharox"],
           ["Spiderling","Decree of Discord"]),
 "Order": (["Ki Barrier","Lacerate","Noxian Emissary","Dragon Form","Kayle Justified","Masa Crashing Thunder","Sacred Protector"],
           ["Disciple of Shen","Kennen Keeper of Balance","Solari Sunhawk","Soulspinner","Ambessa Respected & Feared","Aurok General","Keeper of Law","Horns of the Dragon","Shen Leader of the Kinkou"],
           ["Fallen Feline","Escaped Grayback","Hungry Wolf"],
           ["Glowstone","Reluctant Leader","Shady Spectacles","Decree of Unity"]),
}
TN = {1: "MUST", 2: "GOOD", 3: "SIT", 4: "AVOID"}
TIER = {}
for dom, gs in V7.items():
    for i, g in enumerate(gs, 1):
        for nm in g:
            c = find(nm)
            if c:
                TIER[c["Card Name"]] = i

PIP = {"Fury":"R","Calm":"G","Mind":"B","Body":"O","Chaos":"P","Order":"Y"}


def cost_code(c):
    e, p = c["Energy"].strip(), c["Power"].strip()
    if p in ("", "None", "-"):
        return e
    parts = [x.strip() for x in p.split(",")]
    if all(x not in PIP for x in parts):
        return f"{e}+{len(parts)}"
    u = sorted(set(x for x in parts if x in PIP))
    g = sum(1 for x in parts if x not in PIP)
    t = f"+{g}" if g else ""
    if len(u) == 1:
        return f"{e}{PIP[u[0]] * sum(1 for x in parts if x == u[0])}{t}"
    return f"{e}({'/'.join(PIP[d] for d in u)}){t}"


# Units that cannot be played on turn 1 despite costing 2. EXACT names only —
# substring matching here once caught Punching Poro as well and undercounted.
BARRED_FROM_TURN1 = {"Ol' Poro"}


def is_turn1(c, used_energy):
    return c["type"] == "Unit" and used_energy == 2 and c["card"] not in BARRED_FROM_TURN1


num = lambda v: int(v) if str(v).strip().isdigit() else None
DECKN = {}
for k, v in DECK.items():
    c = find(k)
    assert c, k
    DECKN[c["Card Name"]] = v

recs, miss = [], []
for name, qty, ph in POOL:
    c = find(name)
    if not c:
        miss.append(name); continue
    cn = c["Card Name"]
    recs.append({"card":cn, "code":c["Card Code"], "dom":c["Domain(s)"], "type":c["Type"],
                 "sub":"" if c["Subtype"].strip() in ("None","-","") else c["Subtype"].strip(),
                 "E":num(c["Energy"]), "cc":cost_code(c), "M":num(c["Might"]),
                 "spd":"" if c["Speed"].strip() in ("None","-","") else c["Speed"].strip(),
                 "kw":"" if c["Keywords"].strip() in ("None","-","") else c["Keywords"].strip(),
                 "tier":TN.get(TIER.get(cn), ""), "photo":ph, "qty":qty,
                 "used":DECKN.get(cn, 0)})

DO = {"Fury":1,"Body":2,"Order":3,"Calm":4,"Mind":5,"Chaos":6}
TO = {"Legend":0,"Battlefield":1}
recs.sort(key=lambda r: (TO.get(r["type"], 5), DO.get(r["dom"].split(",")[0].strip(), 9),
                         r["E"] if r["E"] is not None else 99, r["card"]))

wb = Workbook(); ws = wb.active; ws.title = "Pre-Rift Pool S2"
ARI = "Arial"
HF = PatternFill("solid", fgColor="1F3864"); HI = PatternFill("solid", fgColor="7F6000")
IN = PatternFill("solid", fgColor="FFF2CC"); DK = PatternFill("solid", fgColor="D9EAD3")
thin = Side(style="thin", color="BFBFBF"); bd = Border(left=thin, right=thin, top=thin, bottom=thin)

ws["A1"] = "PRE-RIFT POOL \u2014 SESSION 2 \u00b7 read off photos 1\u20137, joined to the Master Card Inventory. Photo 8 looks like the deck you built, so the two right-hand columns are pre-filled from it \u2014 correct anything I got wrong."
ws["A1"].font = Font(name=ARI, size=10, bold=True); ws.merge_cells("A1:P1")
ws["A2"] = "Green rows = in the deck. Your input columns are N and O on the right, plus Notes in P."
ws["A2"].font = Font(name=ARI, size=9, italic=True, color="595959"); ws.merge_cells("A2:P2")

HDR = ["Card","Code","Domain","Type","Subtype","E","Cost code","Might","Speed","Keywords",
       "v7 tier","Photo","Qty opened","In deck?","Copies used","Notes"]
for i, h in enumerate(HDR, 1):
    c = ws.cell(row=4, column=i, value=h)
    c.font = Font(name=ARI, size=10, bold=True, color="FFFFFF")
    c.fill = HI if i >= 14 else HF
    c.alignment = Alignment(horizontal="center", wrap_text=True); c.border = bd

r = 5
for rec in recs:
    vals = [rec["card"], rec["code"], rec["dom"], rec["type"], rec["sub"], rec["E"], rec["cc"],
            rec["M"], rec["spd"], rec["kw"], rec["tier"], rec["photo"], rec["qty"],
            ("Yes" if rec["used"] else "No"), rec["used"], None]
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = Font(name=ARI, size=9); c.border = bd
        if i in (6, 8, 12, 13, 14, 15):
            c.alignment = Alignment(horizontal="center")
        if rec["used"] > 0:
            c.fill = DK
        elif i >= 14:
            c.fill = IN
    r += 1
last = r - 1

SUM = [("Cards in deck (excl. Legend & Battlefields)",
        f'=SUMIF($D$5:$D${last},"Unit",$O$5:$O${last})+SUMIF($D$5:$D${last},"Spell",$O$5:$O${last})+SUMIF($D$5:$D${last},"Gear",$O$5:$O${last})'),
       ("Units", f'=SUMIF($D$5:$D${last},"Unit",$O$5:$O${last})'),
       ("Spells", f'=SUMIF($D$5:$D${last},"Spell",$O$5:$O${last})'),
       ("Gear", f'=SUMIF($D$5:$D${last},"Gear",$O$5:$O${last})'),
       ("Turn-1 units (2-cost; only Ol' Poro excluded, Punching Poro counts)",
        f'=SUMPRODUCT(($D$5:$D${last}="Unit")*($F$5:$F${last}=2)*($A$5:$A${last}<>"Ol\' Poro")*$O$5:$O${last})'),
       ("Units at Might 5+  (target \u22655)",
        f'=SUMPRODUCT(($D$5:$D${last}="Unit")*($H$5:$H${last}>=5)*$O$5:$O${last})'),
       ("Cards at 6+ energy  (max 2)", f'=SUMPRODUCT(($F$5:$F${last}>=6)*$O$5:$O${last})'),
       ("MUST-tier copies played", f'=SUMIF($K$5:$K${last},"MUST",$O$5:$O${last})'),
       ("SIT-tier copies played", f'=SUMIF($K$5:$K${last},"SIT",$O$5:$O${last})')]
for off, (lbl, f) in enumerate(SUM):
    rr = last + 2 + off
    ws.cell(row=rr, column=1, value=lbl).font = Font(name=ARI, size=9, bold=True)
    ws.cell(row=rr, column=2, value=f).font = Font(name=ARI, size=9, bold=True)

for i, w in enumerate([30,14,13,11,13,5,10,7,10,28,8,7,11,9,11,26], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.auto_filter.ref = f"A4:P{last}"

ws2 = wb.create_sheet("Domain compare")
ws2["A1"] = "WHAT EACH DOMAIN OFFERED \u2014 deck-eligible copies in the Session 2 pool"
ws2["A1"].font = Font(name=ARI, size=11, bold=True)
for i, h in enumerate(["Domain","Cards","MUST","GOOD","SIT","AVOID","Turn-1 units","Units M5+","Played?"], 1):
    c = ws2.cell(row=3, column=i, value=h)
    c.font = Font(name=ARI, size=10, bold=True, color="FFFFFF"); c.fill = HF
rr = 4
played = {"Fury","Calm","Chaos"}
for d in ["Fury","Body","Order","Calm","Mind","Chaos"]:
    dr = [x for x in recs if x["type"] in ("Unit","Spell","Gear") and x["dom"].split(",")[0].strip() == d]
    q = lambda f: sum(x["qty"] for x in dr if f(x))
    vals = [d, q(lambda x: True), q(lambda x: x["tier"]=="MUST"), q(lambda x: x["tier"]=="GOOD"),
            q(lambda x: x["tier"]=="SIT"), q(lambda x: x["tier"]=="AVOID"),
            q(lambda x: x["type"]=="Unit" and x["E"]==2 and x["card"] not in BARRED_FROM_TURN1),
            q(lambda x: x["type"]=="Unit" and (x["M"] or 0) >= 5),
            "YES" if d in played else ""]
    for i, v in enumerate(vals, 1):
        ws2.cell(row=rr, column=i, value=v).font = Font(name=ARI, size=10, bold=(i==9 and v=="YES"))
    rr += 1
for i, w in enumerate([12,8,8,8,7,8,14,11,10], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

wb.save("/mnt/user-data/outputs/Pre-Rift_Pool_Session2.xlsx")

deckrows = [x for x in recs if x["used"] and x["type"] in ("Unit","Spell","Gear")]
print("unmatched:", miss)
print("pool copies:", sum(x["qty"] for x in recs),
      "| deck-eligible:", sum(x["qty"] for x in recs if x["type"] in ("Unit","Spell","Gear")))
print("deck size:", sum(x["used"] for x in deckrows),
      "| units", sum(x["used"] for x in deckrows if x["type"]=="Unit"),
      "| spells", sum(x["used"] for x in deckrows if x["type"]=="Spell"),
      "| gear", sum(x["used"] for x in deckrows if x["type"]=="Gear"))
print("turn-1:", sum(x["used"] for x in deckrows if x["type"]=="Unit" and x["E"]==2 and x["card"] not in BARRED_FROM_TURN1),
      "| M5+:", sum(x["used"] for x in deckrows if x["type"]=="Unit" and (x["M"] or 0)>=5),
      "| 6+ energy:", sum(x["used"] for x in deckrows if (x["E"] or 0)>=6))
print("tiers played:", Counter(x["tier"] or "untiered" for x in deckrows for _ in range(x["used"])))
